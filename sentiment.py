import requests
import pandas as pd
import json
import xlsxwriter
from openpyxl import load_workbook


def analyse_comments(rows):
    workbook = load_workbook(filename="bunsen_reviews.xlsx")
    sheet = workbook.active

    count = 0

    headers = {
        'Content-Type': 'application/json',
    }

    params = (
        ('version', '2019-07-12'),
    )
    # responseSrr = ""

    for comment in rows:
        count += 1
        try:
            comment = str(comment.encode('utf-8'))

            data = ' {\n  "text": "' + comment + '",' \
                                                 '\n  "features": {\n    "sentiment": {\n     \n},\n    ' \
                                                 '"keywords": {\n      "emotion": true\n    }\n  }\n}'

            response = requests.post(
                'https://gateway-lon.watsonplatform.net/natural-language-understanding/api/v1/analyze',
                headers=headers, params=params, data=data,
                auth=('apikey', 'BaaDAfn3qpjkCo2Hsm173CLmCKqD-Bv2OIYpnxX4LetC'))

            print(json.dumps(response.text))
            add_senti_analysis(count, json.dumps(response.text), sheet, workbook)
            print("=======================================================================")

            # workbook = xlsxwriter.Workbook(r"C:\Users\eyob\PycharmProjects\bunsen_reviews_c.xlsx")
            # worksheet = workbook.add_worksheet()

            # Start from the first cell.
            # Rows and columns are zero indexed.
            # row = 0
            # column = 0

            # content = json.dumps(response.text)

            # iterating through content list
            # for item in content:
            #     # write operation perform
            #     worksheet.write(column, row, item)
            #
            #     # incrementing the value of row by one
            #     # with each iteratons.
            #     column += 1
            #     workbook.close()

          # responseSrr += "\n" + response.text  #combining json objects

        except Exception as error:
            print(error)
            add_senti_analysis(count, str(error), sheet, workbook)


def add_senti_analysis(count, data, worksheet, book):



    worksheet['E'+str(count)] = data

    book.save("./bunsen_reviews.xlsx")

